from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    '''
    Converts csv to xlsx.
    '''
    cp, xp = Path(csv_path), Path(xlsx_path)
    if not cp.exists(): raise FileNotFoundError(cp)
    if cp.suffix != '.csv': raise ValueError(f'Неверное расширение csv файла {cp.suffix}')
    if xp.suffix != '.xlsx': raise ValueError(f'Неверное расширение xlsx файла {xp.suffix}')


    rows = list(csv.reader(cp.open(encoding="utf-8")))
    if not rows or not rows[0]: raise ValueError("CSV пустой или без заголовка")

    wb, ws = Workbook(), Workbook().active
    ws.title = "Sheet1"
    for r in rows: ws.append(r)

    for i, col in enumerate(ws.iter_cols(values_only=True), 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            8, max(len(str(c or "")) for c in col) + 1
        )
    wb._sheets = [ws]
    wb.save(xlsx_path)
