from pathlib import Path
import csv, json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    '''
    Converts csv to xlsx.
    '''
    cp, xp = Path(csv_path), Path(xlsx_path)
    if not cp.exists(): raise FileNotFoundError(cp)
    if cp.suffix != ".csv" or xp.suffix != ".xlsx":
        raise ValueError("Неверное расширение файла")

    rows = list(csv.reader(cp.open(encoding="utf-8")))
    if not rows or not rows[0]: raise ValueError("CSV пустой или без заголовка")

    wb, ws = Workbook(), Workbook().active
    ws.title = "Sheet1"
    for r in rows: ws.append(r)

    for i, col in enumerate(ws.iter_cols(values_only=True), 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            8, max(len(str(c or "")) for c in col) + 1
        )
    wb._sheets = [ws]
    wb.save(xlsx_path)

def json_to_csv(json_path: str, csv_path: str) -> None:
    jp, cp = Path(json_path), Path(csv_path)
    if not jp.exists(): raise FileNotFoundError(jp)
    if jp.suffix != ".json" or cp.suffix != ".csv":
        raise ValueError(f"Неверное расширение файла {jp, cp}")

    data = json.loads(jp.read_text(encoding="utf-8"))
    if not isinstance(data, list) or not data or not all(isinstance(x, dict) for x in data):
        raise ValueError("Пустой или некорректный JSON")

    keys = list(data[0])
    for d in data[1:]:
        for k in d:
            if k not in keys: keys.append(k)

    with cp.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for d in data: w.writerow({k: d.get(k, "") for k in keys})

def csv_to_json(csv_path: str, json_path: str) -> None:
    cp, jp = Path(csv_path), Path(json_path)
    if not cp.exists(): raise FileNotFoundError(cp)
    if cp.suffix != ".csv" or jp.suffix != ".json":
        raise ValueError("Неверное расширение файла")

    with cp.open(encoding="utf-8") as f:
        r = csv.DictReader(f)
        if not r.fieldnames: raise ValueError("CSV без заголовка или пустой")
        data = [{k: (v or "") for k, v in row.items()} for row in r]
        if not data: raise ValueError("CSV пустой")

    jp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")